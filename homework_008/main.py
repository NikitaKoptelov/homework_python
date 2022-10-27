import wx
import wx.grid as gridlib

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import FORMULAE

class PanelOne(wx.Panel):

    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        txt = wx.TextCtrl(self)


class PanelTwo_filiali(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.currentlySelectedCell = (0, 0)

        self.wb = openpyxl.load_workbook('magazinchik.xlsx')
        self.sheet = self.wb['Филиалы']

        rows = (self.sheet.max_row) + 1
        cols = (self.sheet.max_column) + 1

        self.grid = wx.grid.Grid(self)
        self.grid.CreateGrid(rows, cols)
        for row in range(1, rows):
            for col in range(1, cols):
                bukva = get_column_letter(col)
                ctroka = str(self.sheet[f'{bukva}{row}'].value)
                self.grid.SetCellValue(row - 1, col - 1, ctroka)

        self.grid.Bind(gridlib.EVT_GRID_SELECT_CELL, self.onSingleSelect)
        self.grid.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.onDragSelection)

        selectBtn = wx.Button(self, label="сохранить внесенные изменения")
        selectBtn.Bind(wx.EVT_BUTTON, self.onGetSelection)

        sizer.Add(self.grid, 1, wx.EXPAND)
        sizer.Add(selectBtn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)

    def onDragSelection(self, event):
        """
        Получаем клетки, которые были выбраны посредством зажатия левой
        кнопки мыши и выделения
        """
        if self.grid.GetSelectionBlockTopLeft():
            top_left = self.grid.GetSelectionBlockTopLeft()[0]
            bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
            # self.printSelectedCells(top_left, bottom_right)

    def onGetSelection(self, event):
        """
        Получаем клетки, которые выбраны в данный момент
        """
        cells = self.grid.GetSelectedCells()
        if not cells:
            if self.grid.GetSelectionBlockTopLeft():
                top_left = self.grid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
                self.printSelectedCells(top_left, bottom_right)                      #   вывод данных по нажатию кнопки
            else:
                # print(self.currentlySelectedCell)
                return self.currentlySelectedCell
        else:
            # print(cells)
            return cells

    def onSingleSelect(self, event):
        """
        Получаем выделение одной ячейки посредством клика по ней
        или перемещения между клетками посредством клавиш-стрелок
        """
        # print("You selected Row %s, Col %s" % (event.GetRow(), event.GetCol()))

        self.currentlySelectedCell = (event.GetRow(),
                                      event.GetCol())
        event.Skip()

        # return self.currentlySelectedCell

    def printSelectedCells(self, top_left, bottom_right):
        """
        Основано на коде из: http://ginstrom.com/scribbles/2008/09/07/getting-the-selected-cells-from-a-wxpython-grid/
        """
        cells = []

        rows_start = top_left[0]
        rows_end = bottom_right[0]

        cols_start = top_left[1]
        cols_end = bottom_right[1]

        rows = range(rows_start, rows_end + 1)
        cols = range(cols_start, cols_end + 1)

        cells.extend([(row, col)
                      for row in rows
                      for col in cols])

        print("выделены ячейки: ", cells)


        for cell in cells:
            row, col = cell
            bukva = get_column_letter(col+1)                                  #   вставиить конвертер для координат ЭКСЕЛЯ
            print(f'значение выбраных ячеек - {self.grid.GetCellValue(row, col)} - {row+1} : {bukva}')
            koordinat = str(bukva) + str(row + 1)
            print(f'запись по координатам - {koordinat}', type(koordinat))
            self.wb = openpyxl.load_workbook('magazinchik.xlsx')
            self.sheet = self.wb['Филиалы']
            self.sheet[koordinat]=self.grid.GetCellValue(row, col)
            self.wb.save('magazinchik.xlsx')
                                                           #    в цикле выводится информация по координатам, надо прикрутить запись в ЭКСЕЛЬ по координатам

        def createToolBar(self):                                   # (1) Создание панели инструментов
            toolbar = self.CreateToolBar()
            for each in self.toolbarData():
                self.createSimpleTool(toolbar, *each)
            toolbar.AddSeparator()
            for each in self.toolbarColorData():
                self.createColorTool(toolbar, each)
            toolbar.Realize()

class PanelTwo_tovari(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.currentlySelectedCell = (0, 0)

        self.wb = openpyxl.load_workbook('magazinchik.xlsx')
        self.sheet = self.wb['Товары']

        rows = (self.sheet.max_row) + 1
        cols = (self.sheet.max_column) + 1

        self.grid = wx.grid.Grid(self)
        self.grid.CreateGrid(rows, cols)
        for row in range(1, rows):
            for col in range(1, cols):
                bukva = get_column_letter(col)
                ctroka = str(self.sheet[f'{bukva}{row}'].value)
                self.grid.SetCellValue(row - 1, col - 1, ctroka)

        self.grid.Bind(gridlib.EVT_GRID_SELECT_CELL, self.onSingleSelect)
        self.grid.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.onDragSelection)

        selectBtn = wx.Button(self, label="сохранить внесенные изменения")
        selectBtn.Bind(wx.EVT_BUTTON, self.onGetSelection)

        sizer.Add(self.grid, 1, wx.EXPAND)
        sizer.Add(selectBtn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)

    def onDragSelection(self, event):
        """
        Получаем клетки, которые были выбраны посредством зажатия левой
        кнопки мыши и выделения
        """
        if self.grid.GetSelectionBlockTopLeft():
            top_left = self.grid.GetSelectionBlockTopLeft()[0]
            bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
            # self.printSelectedCells(top_left, bottom_right)

    def onGetSelection(self, event):
        """
        Получаем клетки, которые выбраны в данный момент
        """
        cells = self.grid.GetSelectedCells()
        if not cells:
            if self.grid.GetSelectionBlockTopLeft():
                top_left = self.grid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
                self.printSelectedCells(top_left, bottom_right)                      #   вывод данных по нажатию кнопки
            else:
                # print(self.currentlySelectedCell)
                return self.currentlySelectedCell
        else:
            # print(cells)
            return cells

    def onSingleSelect(self, event):
        """
        Получаем выделение одной ячейки посредством клика по ней
        или перемещения между клетками посредством клавиш-стрелок
        """
        # print("You selected Row %s, Col %s" % (event.GetRow(), event.GetCol()))

        self.currentlySelectedCell = (event.GetRow(),
                                      event.GetCol())
        event.Skip()

        # return self.currentlySelectedCell

    def printSelectedCells(self, top_left, bottom_right):
        """
        Основано на коде из: http://ginstrom.com/scribbles/2008/09/07/getting-the-selected-cells-from-a-wxpython-grid/
        """
        cells = []

        rows_start = top_left[0]
        rows_end = bottom_right[0]

        cols_start = top_left[1]
        cols_end = bottom_right[1]

        rows = range(rows_start, rows_end + 1)
        cols = range(cols_start, cols_end + 1)

        cells.extend([(row, col)
                      for row in rows
                      for col in cols])

        print("выделены ячейки: ", cells)


        for cell in cells:
            row, col = cell
            bukva = get_column_letter(col+1)                                  #   вставиить конвертер для координат ЭКСЕЛЯ
            print(f'значение выбраных ячеек - {self.grid.GetCellValue(row, col)} - {row+1} : {bukva}')
            koordinat = str(bukva) + str(row + 1)
            print(f'запись по координатам - {koordinat}', type(koordinat))
            self.wb = openpyxl.load_workbook('magazinchik.xlsx')
            self.sheet = self.wb['Товары']
            self.sheet[koordinat]=self.grid.GetCellValue(row, col)
            self.wb.save('magazinchik.xlsx')
                                                           #    в цикле выводится информация по координатам, надо прикрутить запись в ЭКСЕЛЬ по координатам

        def createToolBar(self):                                   # (1) Создание панели инструментов
            toolbar = self.CreateToolBar()
            for each in self.toolbarData():
                self.createSimpleTool(toolbar, *each)
            toolbar.AddSeparator()
            for each in self.toolbarColorData():
                self.createColorTool(toolbar, each)
            toolbar.Realize()

class PanelTwo_proizvod(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.currentlySelectedCell = (0, 0)

        self.wb = openpyxl.load_workbook('magazinchik.xlsx')
        self.sheet = self.wb['Производители товаров']

        rows = (self.sheet.max_row) + 1
        cols = (self.sheet.max_column) + 1

        self.grid = wx.grid.Grid(self)
        self.grid.CreateGrid(rows, cols)
        for row in range(1, rows):
            for col in range(1, cols):
                bukva = get_column_letter(col)
                ctroka = str(self.sheet[f'{bukva}{row}'].value)
                self.grid.SetCellValue(row - 1, col - 1, ctroka)

        self.grid.Bind(gridlib.EVT_GRID_SELECT_CELL, self.onSingleSelect)
        self.grid.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.onDragSelection)

        selectBtn = wx.Button(self, label="сохранить внесенные изменения")
        selectBtn.Bind(wx.EVT_BUTTON, self.onGetSelection)

        sizer.Add(self.grid, 1, wx.EXPAND)
        sizer.Add(selectBtn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)

    def onDragSelection(self, event):
        """
        Получаем клетки, которые были выбраны посредством зажатия левой
        кнопки мыши и выделения
        """
        if self.grid.GetSelectionBlockTopLeft():
            top_left = self.grid.GetSelectionBlockTopLeft()[0]
            bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
            # self.printSelectedCells(top_left, bottom_right)

    def onGetSelection(self, event):
        """
        Получаем клетки, которые выбраны в данный момент
        """
        cells = self.grid.GetSelectedCells()
        if not cells:
            if self.grid.GetSelectionBlockTopLeft():
                top_left = self.grid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
                self.printSelectedCells(top_left, bottom_right)                      #   вывод данных по нажатию кнопки
            else:
                # print(self.currentlySelectedCell)
                return self.currentlySelectedCell
        else:
            # print(cells)
            return cells

    def onSingleSelect(self, event):
        """
        Получаем выделение одной ячейки посредством клика по ней
        или перемещения между клетками посредством клавиш-стрелок
        """
        # print("You selected Row %s, Col %s" % (event.GetRow(), event.GetCol()))

        self.currentlySelectedCell = (event.GetRow(),
                                      event.GetCol())
        event.Skip()

        # return self.currentlySelectedCell

    def printSelectedCells(self, top_left, bottom_right):
        """
        Основано на коде из: http://ginstrom.com/scribbles/2008/09/07/getting-the-selected-cells-from-a-wxpython-grid/
        """
        cells = []

        rows_start = top_left[0]
        rows_end = bottom_right[0]

        cols_start = top_left[1]
        cols_end = bottom_right[1]

        rows = range(rows_start, rows_end + 1)
        cols = range(cols_start, cols_end + 1)

        cells.extend([(row, col)
                      for row in rows
                      for col in cols])

        print("выделены ячейки: ", cells)


        for cell in cells:
            row, col = cell
            bukva = get_column_letter(col+1)                                  #   вставиить конвертер для координат ЭКСЕЛЯ
            print(f'значение выбраных ячеек - {self.grid.GetCellValue(row, col)} - {row+1} : {bukva}')
            koordinat = str(bukva) + str(row + 1)
            print(f'запись по координатам - {koordinat}', type(koordinat))
            self.wb = openpyxl.load_workbook('magazinchik.xlsx')
            self.sheet = self.wb['Производители товаров']
            self.sheet[koordinat]=self.grid.GetCellValue(row, col)
            self.wb.save('magazinchik.xlsx')
                                                           #    в цикле выводится информация по координатам, надо прикрутить запись в ЭКСЕЛЬ по координатам

        def createToolBar(self):                                   # (1) Создание панели инструментов
            toolbar = self.CreateToolBar()
            for each in self.toolbarData():
                self.createSimpleTool(toolbar, *each)
            toolbar.AddSeparator()
            for each in self.toolbarColorData():
                self.createColorTool(toolbar, each)
            toolbar.Realize()

class PanelTwo_postavki(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.currentlySelectedCell = (0, 0)

        self.wb = openpyxl.load_workbook('magazinchik.xlsx')
        self.sheet = self.wb['Поставки']

        rows = (self.sheet.max_row) + 1
        cols = (self.sheet.max_column) + 1

        self.grid = wx.grid.Grid(self)
        self.grid.CreateGrid(rows, cols)
        for row in range(1, rows):
            for col in range(1, cols):
                bukva = get_column_letter(col)
                ctroka = str(self.sheet[f'{bukva}{row}'].value)
                self.grid.SetCellValue(row - 1, col - 1, ctroka)

        self.grid.Bind(gridlib.EVT_GRID_SELECT_CELL, self.onSingleSelect)
        self.grid.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.onDragSelection)

        selectBtn = wx.Button(self, label="сохранить внесенные изменения")
        selectBtn.Bind(wx.EVT_BUTTON, self.onGetSelection)

        sizer.Add(self.grid, 1, wx.EXPAND)
        sizer.Add(selectBtn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)

    def onDragSelection(self, event):
        """
        Получаем клетки, которые были выбраны посредством зажатия левой
        кнопки мыши и выделения
        """
        if self.grid.GetSelectionBlockTopLeft():
            top_left = self.grid.GetSelectionBlockTopLeft()[0]
            bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
            # self.printSelectedCells(top_left, bottom_right)

    def onGetSelection(self, event):
        """
        Получаем клетки, которые выбраны в данный момент
        """
        cells = self.grid.GetSelectedCells()
        if not cells:
            if self.grid.GetSelectionBlockTopLeft():
                top_left = self.grid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
                self.printSelectedCells(top_left, bottom_right)                      #   вывод данных по нажатию кнопки
            else:
                # print(self.currentlySelectedCell)
                return self.currentlySelectedCell
        else:
            # print(cells)
            return cells

    def onSingleSelect(self, event):
        """
        Получаем выделение одной ячейки посредством клика по ней
        или перемещения между клетками посредством клавиш-стрелок
        """
        # print("You selected Row %s, Col %s" % (event.GetRow(), event.GetCol()))

        self.currentlySelectedCell = (event.GetRow(),
                                      event.GetCol())
        event.Skip()

        # return self.currentlySelectedCell

    def printSelectedCells(self, top_left, bottom_right):
        """
        Основано на коде из: http://ginstrom.com/scribbles/2008/09/07/getting-the-selected-cells-from-a-wxpython-grid/
        """
        cells = []

        rows_start = top_left[0]
        rows_end = bottom_right[0]

        cols_start = top_left[1]
        cols_end = bottom_right[1]

        rows = range(rows_start, rows_end + 1)
        cols = range(cols_start, cols_end + 1)

        cells.extend([(row, col)
                      for row in rows
                      for col in cols])

        print("выделены ячейки: ", cells)


        for cell in cells:
            row, col = cell
            bukva = get_column_letter(col+1)                                  #   вставиить конвертер для координат ЭКСЕЛЯ
            print(f'значение выбраных ячеек - {self.grid.GetCellValue(row, col)} - {row+1} : {bukva}')
            koordinat = str(bukva) + str(row + 1)
            print(f'запись по координатам - {koordinat}', type(koordinat))
            self.wb = openpyxl.load_workbook('magazinchik.xlsx')
            self.sheet = self.wb['Поставки']
            self.sheet[koordinat]=self.grid.GetCellValue(row, col)
            self.wb.save('magazinchik.xlsx')
                                                           #    в цикле выводится информация по координатам, надо прикрутить запись в ЭКСЕЛЬ по координатам

        def createToolBar(self):                                   # (1) Создание панели инструментов
            toolbar = self.CreateToolBar()
            for each in self.toolbarData():
                self.createSimpleTool(toolbar, *each)
            toolbar.AddSeparator()
            for each in self.toolbarColorData():
                self.createColorTool(toolbar, each)
            toolbar.Realize()

class PanelTwo_pokupki(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.currentlySelectedCell = (0, 0)

        self.wb = openpyxl.load_workbook('magazinchik.xlsx')
        self.sheet = self.wb['Покупки']

        rows = (self.sheet.max_row) + 1
        cols = (self.sheet.max_column) + 1

        self.grid = wx.grid.Grid(self)
        self.grid.CreateGrid(rows, cols)
        for row in range(1, rows):
            for col in range(1, cols):
                bukva = get_column_letter(col)
                ctroka = str(self.sheet[f'{bukva}{row}'].value)
                self.grid.SetCellValue(row - 1, col - 1, ctroka)

        self.grid.Bind(gridlib.EVT_GRID_SELECT_CELL, self.onSingleSelect)
        self.grid.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.onDragSelection)

        selectBtn = wx.Button(self, label="сохранить внесенные изменения")
        selectBtn.Bind(wx.EVT_BUTTON, self.onGetSelection)

        sizer.Add(self.grid, 1, wx.EXPAND)
        sizer.Add(selectBtn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)

    def onDragSelection(self, event):
        """
        Получаем клетки, которые были выбраны посредством зажатия левой
        кнопки мыши и выделения
        """
        if self.grid.GetSelectionBlockTopLeft():
            top_left = self.grid.GetSelectionBlockTopLeft()[0]
            bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
            # self.printSelectedCells(top_left, bottom_right)

    def onGetSelection(self, event):
        """
        Получаем клетки, которые выбраны в данный момент
        """
        cells = self.grid.GetSelectedCells()
        if not cells:
            if self.grid.GetSelectionBlockTopLeft():
                top_left = self.grid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
                self.printSelectedCells(top_left, bottom_right)                      #   вывод данных по нажатию кнопки
            else:
                # print(self.currentlySelectedCell)
                return self.currentlySelectedCell
        else:
            # print(cells)
            return cells

    def onSingleSelect(self, event):
        """
        Получаем выделение одной ячейки посредством клика по ней
        или перемещения между клетками посредством клавиш-стрелок
        """
        # print("You selected Row %s, Col %s" % (event.GetRow(), event.GetCol()))

        self.currentlySelectedCell = (event.GetRow(),
                                      event.GetCol())
        event.Skip()

        # return self.currentlySelectedCell

    def printSelectedCells(self, top_left, bottom_right):
        """
        Основано на коде из: http://ginstrom.com/scribbles/2008/09/07/getting-the-selected-cells-from-a-wxpython-grid/
        """
        cells = []

        rows_start = top_left[0]
        rows_end = bottom_right[0]

        cols_start = top_left[1]
        cols_end = bottom_right[1]

        rows = range(rows_start, rows_end + 1)
        cols = range(cols_start, cols_end + 1)

        cells.extend([(row, col)
                      for row in rows
                      for col in cols])

        print("выделены ячейки: ", cells)


        for cell in cells:
            row, col = cell
            bukva = get_column_letter(col+1)                                  #   вставиить конвертер для координат ЭКСЕЛЯ
            print(f'значение выбраных ячеек - {self.grid.GetCellValue(row, col)} - {row+1} : {bukva}')
            koordinat = str(bukva) + str(row + 1)
            print(f'запись по координатам - {koordinat}', type(koordinat))
            self.wb = openpyxl.load_workbook('magazinchik.xlsx')
            self.sheet = self.wb['Покупки']
            self.sheet[koordinat]=self.grid.GetCellValue(row, col)
            self.wb.save('magazinchik.xlsx')
                                                           #    в цикле выводится информация по координатам, надо прикрутить запись в ЭКСЕЛЬ по координатам

        def createToolBar(self):                                   # (1) Создание панели инструментов
            toolbar = self.CreateToolBar()
            for each in self.toolbarData():
                self.createSimpleTool(toolbar, *each)
            toolbar.AddSeparator()
            for each in self.toolbarColorData():
                self.createColorTool(toolbar, each)
            toolbar.Realize()

class PanelTwo_klienti(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.currentlySelectedCell = (0, 0)

        self.wb = openpyxl.load_workbook('magazinchik.xlsx')
        self.sheet = self.wb['Клиенты']

        rows = (self.sheet.max_row) + 1
        cols = (self.sheet.max_column) + 1

        self.grid = wx.grid.Grid(self)
        self.grid.CreateGrid(rows, cols)
        for row in range(1, rows):
            for col in range(1, cols):
                bukva = get_column_letter(col)
                ctroka = str(self.sheet[f'{bukva}{row}'].value)
                self.grid.SetCellValue(row - 1, col - 1, ctroka)

        self.grid.Bind(gridlib.EVT_GRID_SELECT_CELL, self.onSingleSelect)
        self.grid.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.onDragSelection)

        selectBtn = wx.Button(self, label="сохранить внесенные изменения")
        selectBtn.Bind(wx.EVT_BUTTON, self.onGetSelection)

        sizer.Add(self.grid, 1, wx.EXPAND)
        sizer.Add(selectBtn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)

    def onDragSelection(self, event):
        """
        Получаем клетки, которые были выбраны посредством зажатия левой
        кнопки мыши и выделения
        """
        if self.grid.GetSelectionBlockTopLeft():
            top_left = self.grid.GetSelectionBlockTopLeft()[0]
            bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
            # self.printSelectedCells(top_left, bottom_right)

    def onGetSelection(self, event):
        """
        Получаем клетки, которые выбраны в данный момент
        """
        cells = self.grid.GetSelectedCells()
        if not cells:
            if self.grid.GetSelectionBlockTopLeft():
                top_left = self.grid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
                self.printSelectedCells(top_left, bottom_right)                      #   вывод данных по нажатию кнопки
            else:
                # print(self.currentlySelectedCell)
                return self.currentlySelectedCell
        else:
            # print(cells)
            return cells

    def onSingleSelect(self, event):
        """
        Получаем выделение одной ячейки посредством клика по ней
        или перемещения между клетками посредством клавиш-стрелок
        """
        # print("You selected Row %s, Col %s" % (event.GetRow(), event.GetCol()))

        self.currentlySelectedCell = (event.GetRow(),
                                      event.GetCol())
        event.Skip()

        # return self.currentlySelectedCell

    def printSelectedCells(self, top_left, bottom_right):
        """
        Основано на коде из: http://ginstrom.com/scribbles/2008/09/07/getting-the-selected-cells-from-a-wxpython-grid/
        """
        cells = []

        rows_start = top_left[0]
        rows_end = bottom_right[0]

        cols_start = top_left[1]
        cols_end = bottom_right[1]

        rows = range(rows_start, rows_end + 1)
        cols = range(cols_start, cols_end + 1)

        cells.extend([(row, col)
                      for row in rows
                      for col in cols])

        print("выделены ячейки: ", cells)


        for cell in cells:
            row, col = cell
            bukva = get_column_letter(col+1)                                  #   вставиить конвертер для координат ЭКСЕЛЯ
            print(f'значение выбраных ячеек - {self.grid.GetCellValue(row, col)} - {row+1} : {bukva}')
            koordinat = str(bukva) + str(row + 1)
            print(f'запись по координатам - {koordinat}', type(koordinat))
            self.wb = openpyxl.load_workbook('magazinchik.xlsx')
            self.sheet = self.wb['Клиенты']
            self.sheet[koordinat]=self.grid.GetCellValue(row, col)
            self.wb.save('magazinchik.xlsx')
                                                           #    в цикле выводится информация по координатам, надо прикрутить запись в ЭКСЕЛЬ по координатам

        def createToolBar(self):                                   # (1) Создание панели инструментов
            toolbar = self.CreateToolBar()
            for each in self.toolbarData():
                self.createSimpleTool(toolbar, *each)
            toolbar.AddSeparator()
            for each in self.toolbarColorData():
                self.createColorTool(toolbar, each)
            toolbar.Realize()

class PanelTwo_kategorii(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.currentlySelectedCell = (0, 0)

        self.wb = openpyxl.load_workbook('magazinchik.xlsx')
        self.sheet = self.wb['Категории']

        rows = (self.sheet.max_row) + 1
        cols = (self.sheet.max_column) + 1

        self.grid = wx.grid.Grid(self)
        self.grid.CreateGrid(rows, cols)
        for row in range(1, rows):
            for col in range(1, cols):
                bukva = get_column_letter(col)
                ctroka = str(self.sheet[f'{bukva}{row}'].value)
                self.grid.SetCellValue(row - 1, col - 1, ctroka)

        self.grid.Bind(gridlib.EVT_GRID_SELECT_CELL, self.onSingleSelect)
        self.grid.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.onDragSelection)

        selectBtn = wx.Button(self, label="сохранить внесенные изменения")
        selectBtn.Bind(wx.EVT_BUTTON, self.onGetSelection)

        sizer.Add(self.grid, 1, wx.EXPAND)
        sizer.Add(selectBtn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)

    def onDragSelection(self, event):
        """
        Получаем клетки, которые были выбраны посредством зажатия левой
        кнопки мыши и выделения
        """
        if self.grid.GetSelectionBlockTopLeft():
            top_left = self.grid.GetSelectionBlockTopLeft()[0]
            bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
            # self.printSelectedCells(top_left, bottom_right)

    def onGetSelection(self, event):
        """
        Получаем клетки, которые выбраны в данный момент
        """
        cells = self.grid.GetSelectedCells()
        if not cells:
            if self.grid.GetSelectionBlockTopLeft():
                top_left = self.grid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
                self.printSelectedCells(top_left, bottom_right)                      #   вывод данных по нажатию кнопки
            else:
                # print(self.currentlySelectedCell)
                return self.currentlySelectedCell
        else:
            # print(cells)
            return cells

    def onSingleSelect(self, event):
        """
        Получаем выделение одной ячейки посредством клика по ней
        или перемещения между клетками посредством клавиш-стрелок
        """
        # print("You selected Row %s, Col %s" % (event.GetRow(), event.GetCol()))

        self.currentlySelectedCell = (event.GetRow(),
                                      event.GetCol())
        event.Skip()

        # return self.currentlySelectedCell

    def printSelectedCells(self, top_left, bottom_right):
        """
        Основано на коде из: http://ginstrom.com/scribbles/2008/09/07/getting-the-selected-cells-from-a-wxpython-grid/
        """
        cells = []

        rows_start = top_left[0]
        rows_end = bottom_right[0]

        cols_start = top_left[1]
        cols_end = bottom_right[1]

        rows = range(rows_start, rows_end + 1)
        cols = range(cols_start, cols_end + 1)

        cells.extend([(row, col)
                      for row in rows
                      for col in cols])

        print("выделены ячейки: ", cells)


        for cell in cells:
            row, col = cell
            bukva = get_column_letter(col+1)                                  #   вставиить конвертер для координат ЭКСЕЛЯ
            print(f'значение выбраных ячеек - {self.grid.GetCellValue(row, col)} - {row+1} : {bukva}')
            koordinat = str(bukva) + str(row + 1)
            print(f'запись по координатам - {koordinat}', type(koordinat))
            self.wb = openpyxl.load_workbook('magazinchik.xlsx')
            self.sheet = self.wb['Категории']
            self.sheet[koordinat]=self.grid.GetCellValue(row, col)
            self.wb.save('magazinchik.xlsx')
                                                           #    в цикле выводится информация по координатам, надо прикрутить запись в ЭКСЕЛЬ по координатам

        def createToolBar(self):                                   # (1) Создание панели инструментов
            toolbar = self.CreateToolBar()
            for each in self.toolbarData():
                self.createSimpleTool(toolbar, *each)
            toolbar.AddSeparator()
            for each in self.toolbarColorData():
                self.createColorTool(toolbar, each)
            toolbar.Realize()

class PanelTwo_izmen(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.currentlySelectedCell = (0, 0)

        self.wb = openpyxl.load_workbook('magazinchik.xlsx')
        self.sheet = self.wb['Изменение цены на товары']

        rows = (self.sheet.max_row) + 1
        cols = (self.sheet.max_column) + 1

        self.grid = wx.grid.Grid(self)
        self.grid.CreateGrid(rows, cols)
        for row in range(1, rows):
            for col in range(1, cols):
                bukva = get_column_letter(col)
                ctroka = str(self.sheet[f'{bukva}{row}'].value)
                self.grid.SetCellValue(row - 1, col - 1, ctroka)

        self.grid.Bind(gridlib.EVT_GRID_SELECT_CELL, self.onSingleSelect)
        self.grid.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.onDragSelection)

        selectBtn = wx.Button(self, label="сохранить внесенные изменения")
        selectBtn.Bind(wx.EVT_BUTTON, self.onGetSelection)

        sizer.Add(self.grid, 1, wx.EXPAND)
        sizer.Add(selectBtn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)

    def onDragSelection(self, event):
        """
        Получаем клетки, которые были выбраны посредством зажатия левой
        кнопки мыши и выделения
        """
        if self.grid.GetSelectionBlockTopLeft():
            top_left = self.grid.GetSelectionBlockTopLeft()[0]
            bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
            # self.printSelectedCells(top_left, bottom_right)

    def onGetSelection(self, event):
        """
        Получаем клетки, которые выбраны в данный момент
        """
        cells = self.grid.GetSelectedCells()
        if not cells:
            if self.grid.GetSelectionBlockTopLeft():
                top_left = self.grid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
                self.printSelectedCells(top_left, bottom_right)                      #   вывод данных по нажатию кнопки
            else:
                # print(self.currentlySelectedCell)
                return self.currentlySelectedCell
        else:
            # print(cells)
            return cells

    def onSingleSelect(self, event):
        """
        Получаем выделение одной ячейки посредством клика по ней
        или перемещения между клетками посредством клавиш-стрелок
        """
        # print("You selected Row %s, Col %s" % (event.GetRow(), event.GetCol()))

        self.currentlySelectedCell = (event.GetRow(),
                                      event.GetCol())
        event.Skip()

        # return self.currentlySelectedCell

    def printSelectedCells(self, top_left, bottom_right):
        """
        Основано на коде из: http://ginstrom.com/scribbles/2008/09/07/getting-the-selected-cells-from-a-wxpython-grid/
        """
        cells = []

        rows_start = top_left[0]
        rows_end = bottom_right[0]

        cols_start = top_left[1]
        cols_end = bottom_right[1]

        rows = range(rows_start, rows_end + 1)
        cols = range(cols_start, cols_end + 1)

        cells.extend([(row, col)
                      for row in rows
                      for col in cols])

        print("выделены ячейки: ", cells)


        for cell in cells:
            row, col = cell
            bukva = get_column_letter(col+1)                                  #   вставиить конвертер для координат ЭКСЕЛЯ
            print(f'значение выбраных ячеек - {self.grid.GetCellValue(row, col)} - {row+1} : {bukva}')
            koordinat = str(bukva) + str(row + 1)
            print(f'запись по координатам - {koordinat}', type(koordinat))
            self.wb = openpyxl.load_workbook('magazinchik.xlsx')
            self.sheet = self.wb['Изменение цены на товары']
            self.sheet[koordinat]=self.grid.GetCellValue(row, col)
            self.wb.save('magazinchik.xlsx')
                                                           #    в цикле выводится информация по координатам, надо прикрутить запись в ЭКСЕЛЬ по координатам

        def createToolBar(self):                                   # (1) Создание панели инструментов
            toolbar = self.CreateToolBar()
            for each in self.toolbarData():
                self.createSimpleTool(toolbar, *each)
            toolbar.AddSeparator()
            for each in self.toolbarColorData():
                self.createColorTool(toolbar, each)
            toolbar.Realize()

class PanelTwo_zapisi(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent=parent)
        sizer = wx.BoxSizer(wx.VERTICAL)

        self.currentlySelectedCell = (0, 0)

        self.wb = openpyxl.load_workbook('magazinchik.xlsx')
        self.sheet = self.wb['Запись в счете']

        rows = (self.sheet.max_row) + 1
        cols = (self.sheet.max_column) + 1

        self.grid = wx.grid.Grid(self)
        self.grid.CreateGrid(rows, cols)
        for row in range(1, rows):
            for col in range(1, cols):
                bukva = get_column_letter(col)
                ctroka = str(self.sheet[f'{bukva}{row}'].value)
                self.grid.SetCellValue(row - 1, col - 1, ctroka)

        self.grid.Bind(gridlib.EVT_GRID_SELECT_CELL, self.onSingleSelect)
        self.grid.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.onDragSelection)

        selectBtn = wx.Button(self, label="сохранить внесенные изменения")
        selectBtn.Bind(wx.EVT_BUTTON, self.onGetSelection)

        sizer.Add(self.grid, 1, wx.EXPAND)
        sizer.Add(selectBtn, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)

    def onDragSelection(self, event):
        """
        Получаем клетки, которые были выбраны посредством зажатия левой
        кнопки мыши и выделения
        """
        if self.grid.GetSelectionBlockTopLeft():
            top_left = self.grid.GetSelectionBlockTopLeft()[0]
            bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
            # self.printSelectedCells(top_left, bottom_right)

    def onGetSelection(self, event):
        """
        Получаем клетки, которые выбраны в данный момент
        """
        cells = self.grid.GetSelectedCells()
        if not cells:
            if self.grid.GetSelectionBlockTopLeft():
                top_left = self.grid.GetSelectionBlockTopLeft()[0]
                bottom_right = self.grid.GetSelectionBlockBottomRight()[0]
                self.printSelectedCells(top_left, bottom_right)                      #   вывод данных по нажатию кнопки
            else:
                # print(self.currentlySelectedCell)
                return self.currentlySelectedCell
        else:
            # print(cells)
            return cells

    def onSingleSelect(self, event):
        """
        Получаем выделение одной ячейки посредством клика по ней
        или перемещения между клетками посредством клавиш-стрелок
        """
        # print("You selected Row %s, Col %s" % (event.GetRow(), event.GetCol()))

        self.currentlySelectedCell = (event.GetRow(),
                                      event.GetCol())
        event.Skip()

        # return self.currentlySelectedCell

    def printSelectedCells(self, top_left, bottom_right):
        """
        Основано на коде из: http://ginstrom.com/scribbles/2008/09/07/getting-the-selected-cells-from-a-wxpython-grid/
        """
        cells = []

        rows_start = top_left[0]
        rows_end = bottom_right[0]

        cols_start = top_left[1]
        cols_end = bottom_right[1]

        rows = range(rows_start, rows_end + 1)
        cols = range(cols_start, cols_end + 1)

        cells.extend([(row, col)
                      for row in rows
                      for col in cols])

        print("выделены ячейки: ", cells)


        for cell in cells:
            row, col = cell
            bukva = get_column_letter(col+1)                                  #   вставиить конвертер для координат ЭКСЕЛЯ
            print(f'значение выбраных ячеек - {self.grid.GetCellValue(row, col)} - {row+1} : {bukva}')
            koordinat = str(bukva) + str(row + 1)
            print(f'запись по координатам - {koordinat}', type(koordinat))
            self.wb = openpyxl.load_workbook('magazinchik.xlsx')
            self.sheet = self.wb['Запись в счете']
            self.sheet[koordinat]=self.grid.GetCellValue(row, col)
            self.wb.save('magazinchik.xlsx')
                                                           #    в цикле выводится информация по координатам, надо прикрутить запись в ЭКСЕЛЬ по координатам

        def createToolBar(self):                                   # (1) Создание панели инструментов
            toolbar = self.CreateToolBar()
            for each in self.toolbarData():
                self.createSimpleTool(toolbar, *each)
            toolbar.AddSeparator()
            for each in self.toolbarColorData():
                self.createColorTool(toolbar, each)
            toolbar.Realize()

class MyForm(wx.Frame):
    def __init__(self):
        wx.Frame.__init__(self, None, wx.ID_ANY, "Моя первая программа")

        # self.name_tab = ''
        self.panel_glav = PanelOne(self)
        self.panel_tab_1 = PanelTwo_filiali(self)
        self.panel_tab_2 = PanelTwo_tovari(self)
        self.panel_tab_3 = PanelTwo_proizvod(self)
        self.panel_tab_4 = PanelTwo_postavki(self)
        self.panel_tab_5 = PanelTwo_pokupki(self)
        self.panel_tab_6 = PanelTwo_klienti(self)
        self.panel_tab_7 = PanelTwo_kategorii(self)
        self.panel_tab_8 = PanelTwo_izmen(self)
        self.panel_tab_9 = PanelTwo_zapisi(self)
        self.panel_glav.Show()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Hide()



        self.sizer = wx.BoxSizer(wx.VERTICAL)
        self.sizer.Add(self.panel_glav, 1, wx.EXPAND)
        self.sizer.Add(self.panel_tab_1, 1, wx.EXPAND)
        self.sizer.Add(self.panel_tab_2, 1, wx.EXPAND)
        self.sizer.Add(self.panel_tab_3, 1, wx.EXPAND)
        self.sizer.Add(self.panel_tab_4, 1, wx.EXPAND)
        self.sizer.Add(self.panel_tab_5, 1, wx.EXPAND)
        self.sizer.Add(self.panel_tab_6, 1, wx.EXPAND)
        self.sizer.Add(self.panel_tab_7, 1, wx.EXPAND)
        self.sizer.Add(self.panel_tab_8, 1, wx.EXPAND)
        self.sizer.Add(self.panel_tab_9, 1, wx.EXPAND)
        self.SetSizer(self.sizer)

        # self.sizer = wx.BoxSizer()                                          #    без этих строк не появляется меню
        # self.SetSizer(self.sizer)                                                      #    происходит создание места для создания меню

        menubar = wx.MenuBar()
        fileMenu = wx.Menu()

        open_panel_g = fileMenu.Append(wx.ID_ANY, 'Главное окно')
        self.Bind(wx.EVT_MENU, self.on_panel_g, open_panel_g)
        open_panel_1 = fileMenu.Append(wx.ID_ANY, 'Таблица - Филиалы')
        self.Bind(wx.EVT_MENU, self.on_panel_1, open_panel_1)
        open_panel_2 = fileMenu.Append(wx.ID_ANY, 'Таблица - Товары')
        self.Bind(wx.EVT_MENU, self.on_panel_2, open_panel_2)
        open_panel_3 = fileMenu.Append(wx.ID_ANY, 'Таблица - Производители')
        self.Bind(wx.EVT_MENU, self.on_panel_3, open_panel_3)
        open_panel_4 = fileMenu.Append(wx.ID_ANY, 'Таблица - Поставки')
        self.Bind(wx.EVT_MENU, self.on_panel_4, open_panel_4)
        open_panel_5 = fileMenu.Append(wx.ID_ANY, 'Таблица - Покупки')
        self.Bind(wx.EVT_MENU, self.on_panel_5, open_panel_5)
        open_panel_6 = fileMenu.Append(wx.ID_ANY, 'Таблица - Клиенты')
        self.Bind(wx.EVT_MENU, self.on_panel_6, open_panel_6)
        open_panel_7 = fileMenu.Append(wx.ID_ANY, 'Таблица - Категории')
        self.Bind(wx.EVT_MENU, self.on_panel_7, open_panel_7)
        open_panel_8 = fileMenu.Append(wx.ID_ANY, 'Таблица - Изменение')
        self.Bind(wx.EVT_MENU, self.on_panel_8, open_panel_8)
        open_panel_9 = fileMenu.Append(wx.ID_ANY, 'Таблица - Запись')
        self.Bind(wx.EVT_MENU, self.on_panel_9, open_panel_9)

        fileMenu.AppendSeparator()                                                       # разделительная линия в нутри меню
        item = fileMenu.Append(wx.ID_EXIT, "Выход", "Выход из приложения")               #  сокращенный вид добавления пунктов меню
        menubar.Append(fileMenu, '&меню')
        self.SetMenuBar(menubar)                                                      #   вызов самого меню

        self.Bind(wx.EVT_MENU, self.onQuit, item)                                    # строка вызова функции для обработки нажатия кнопки

    def onQuit(self, event):                                                       #   функция выхода и закрытия приложения
        self.Close()

    def on_panel_g(self, event):
        self.panel_glav.Show()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Hide()

    def on_panel_1(self, event):
        self.panel_glav.Hide()
        self.panel_tab_1.Show()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Hide()

    def on_panel_2(self, event):
        self.panel_glav.Hide()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Show()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Hide()

    def on_panel_3(self, event):
        self.panel_glav.Hide()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Show()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Hide()

    def on_panel_4(self, event):
        self.panel_glav.Hide()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Show()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Hide()

    def on_panel_5(self, event):
        self.panel_glav.Hide()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Show()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Hide()

    def on_panel_6(self, event):
        self.panel_glav.Hide()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Show()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Hide()

    def on_panel_7(self, event):
        self.panel_glav.Hide()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Show()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Hide()

    def on_panel_8(self, event):
        self.panel_glav.Hide()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Show()
        self.panel_tab_9.Hide()

    def on_panel_9(self, event):
        self.panel_glav.Hide()
        self.panel_tab_1.Hide()
        self.panel_tab_2.Hide()
        self.panel_tab_3.Hide()
        self.panel_tab_4.Hide()
        self.panel_tab_5.Hide()
        self.panel_tab_6.Hide()
        self.panel_tab_7.Hide()
        self.panel_tab_8.Hide()
        self.panel_tab_9.Show()


app = wx.App()
frame = MyForm()
frame.Show()
app.MainLoop()


