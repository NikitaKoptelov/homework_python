import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import FORMULAE

# читаем excel-файл
wb = openpyxl.load_workbook('D:/lesson_Python/homework_python/homework_008/magazinchik.xlsx')

tast_zapisi = ''

def init():
    global wb
    global tast_zapisi
    # global sheet
    # global rows
    # global cols

def schit_tab_Филиалы():
    sheet = wb['Филиалы']
    rows = sheet.max_row
    cols = sheet.max_column
    res = []
    res_2 = []
    index_cols = 0
    index_rows = 0
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            bukva = get_column_letter(j)
            ctroka = sheet[f'{bukva}{i}'].value
            res.append(ctroka)
            index_cols+=1
        res_2.append(res)
        res = []
        index_rows+=1
        index_cols=0
    return res_2

def schit_tab_Товары():
    sheet = wb['Товары']
    rows = sheet.max_row
    cols = sheet.max_column
    res = []
    res_2 = []
    index_cols = 0
    index_rows = 0
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            bukva = get_column_letter(j)
            ctroka = sheet[f'{bukva}{i}'].value
            res.append(ctroka)
            index_cols+=1
        res_2.append(res)
        res = []
        index_rows+=1
        index_cols=0
    return res_2

def schit_tab_Производители_товаров():
    sheet = wb['Производители товаров']
    rows = sheet.max_row
    cols = sheet.max_column
    res = []
    res_2 = []
    index_cols = 0
    index_rows = 0
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            bukva = get_column_letter(j)
            ctroka = sheet[f'{bukva}{i}'].value
            res.append(ctroka)
            index_cols+=1
        res_2.append(res)
        res = []
        index_rows+=1
        index_cols=0
    return res_2

def schit_tab_Поставки():
    sheet = wb['Поставки']
    rows = sheet.max_row
    cols = sheet.max_column
    res = []
    res_2 = []
    index_cols = 0
    index_rows = 0
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            bukva = get_column_letter(j)
            ctroka = sheet[f'{bukva}{i}'].value
            res.append(ctroka)
            index_cols+=1
        res_2.append(res)
        res = []
        index_rows+=1
        index_cols=0
    return res_2

def schit_tab_Покупки():
    sheet = wb['Покупки']
    rows = sheet.max_row
    cols = sheet.max_column
    res = []
    res_2 = []
    index_cols = 0
    index_rows = 0
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            bukva = get_column_letter(j)
            ctroka = sheet[f'{bukva}{i}'].value
            res.append(ctroka)
            index_cols+=1
        res_2.append(res)
        res = []
        index_rows+=1
        index_cols=0
    return res_2

def schit_tab_Клиенты():
    sheet = wb['Клиенты']
    rows = sheet.max_row
    cols = sheet.max_column
    res = []
    res_2 = []
    index_cols = 0
    index_rows = 0
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            bukva = get_column_letter(j)
            ctroka = sheet[f'{bukva}{i}'].value
            res.append(ctroka)
            index_cols+=1
        res_2.append(res)
        res = []
        index_rows+=1
        index_cols=0
    return res_2

def schit_tab_Категории():
    sheet = wb['Категории']
    rows = sheet.max_row
    cols = sheet.max_column
    res = []
    res_2 = []
    index_cols = 0
    index_rows = 0
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            bukva = get_column_letter(j)
            ctroka = sheet[f'{bukva}{i}'].value
            res.append(ctroka)
            index_cols+=1
        res_2.append(res)
        res = []
        index_rows+=1
        index_cols=0
    return res_2

def schit_tab_Изменение_цены_на_товары():
    sheet = wb['Изменение цены на товары']
    rows = sheet.max_row
    cols = sheet.max_column
    res = []
    res_2 = []
    index_cols = 0
    index_rows = 0
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            bukva = get_column_letter(j)
            ctroka = sheet[f'{bukva}{i}'].value
            res.append(ctroka)
            index_cols+=1
        res_2.append(res)
        res = []
        index_rows+=1
        index_cols=0
    return res_2

def schit_tab_Запись_в_счете():
    sheet = wb['Запись в счете']
    rows = sheet.max_row
    cols = sheet.max_column
    res = []
    res_2 = []
    index_cols = 0
    index_rows = 0
    for i in range(1, rows+1):
        for j in range(1, cols+1):
            bukva = get_column_letter(j)
            ctroka = sheet[f'{bukva}{i}'].value
            res.append(ctroka)
            index_cols+=1
        res_2.append(res)
        res = []
        index_rows+=1
        index_cols=0
    return res_2

def wread_file():
    with open("D:/lesson_Python/homework_python/homework_008/text.txt", "a") as write_file:
        write_file.write(tast_zapisi)

